# views.py
from django.shortcuts import render, redirect
from .forms import AsistenciaForm
from .models import Asistencia, Trabajador
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.utils.decorators import method_decorator
from django.contrib.auth.views import LoginView
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import timedelta
from .models import Trabajador, Asistencia, FinalAsistencia
import datetime
from openpyxl.styles import Alignment

def generar_excel_asistencias(request):
    # Crear un nuevo libro de trabajo
    wb = Workbook()
    # Eliminar la hoja por defecto
    wb.remove(wb.active)

    # Obtener todos los trabajadores
    trabajadores = Trabajador.objects.all()

    for trabajador in trabajadores:
        # Crear una hoja para cada trabajador
        hoja = wb.create_sheet(title=trabajador.name)

        # Encabezados
        hoja['A1'] = 'Fecha'
        hoja['B1'] = 'Ingreso'
        hoja['C1'] = 'Salida'
        hoja['D1'] = 'Refrigerio y/o almuerzo'
        hoja['E1'] = 'Hora Total'

        # Obtener las FinalAsistencia del 
        Asistencia.crear_superasistencias(trabajador=trabajador)
        final_asistencias = FinalAsistencia.objects.filter(trabajador=trabajador).order_by('entrada__fecha', 'entrada__hora')
        
        fila = 2  # Comienza en la fila 2
        for final_asistencia in final_asistencias:
            fecha = final_asistencia.entrada.fecha
            ingreso = final_asistencia.entrada.hora
            salida = final_asistencia.salida.hora

            ingreso_timedelta = timedelta(hours=ingreso.hour, minutes=ingreso.minute, seconds=ingreso.second)
            salida_timedelta = timedelta(hours=salida.hour, minutes=salida.minute, seconds=salida.second)
            refrigerio = timedelta(minutes=0)
            hora_total = salida_timedelta - ingreso_timedelta - refrigerio

            # Evitar días negativos y formatear tiempo total
            horas, resto = divmod(hora_total.total_seconds(), 3600)
            minutos, segundos = divmod(resto, 60)
            hora_total = f"{int(horas):02}:{int(minutos):02}:{int(segundos):02}"
            

            hoja[f'A{fila}'] = fecha
            hoja[f'B{fila}'] = ingreso
            hoja[f'C{fila}'] = salida
            hoja[f'D{fila}'] = str(refrigerio)
            hoja[f'E{fila}'] = str(hora_total)

            fila += 1

        # Ajustar ancho de las columnas
        for col in range(1, 6):
            hoja.column_dimensions[get_column_letter(col)].width = 20

        # Centrar el texto en todas las celdas
        for row in hoja.iter_rows(min_row=1, max_row=fila - 1, min_col=1, max_col=5):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')

    # Guardar como respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=asistencias_trabajadores.xlsx'
    wb.save(response)
    return response

def exportar_asistencias_a_excel(request):
    # Crea un libro de trabajo de Excel
    wb = Workbook()
    wb.remove(wb.active)  # Eliminar la hoja predeterminada

    # Obtener todos los trabajadores
    trabajadores = Trabajador.objects.all()

    for trabajador in trabajadores:
        # Crear una hoja para cada trabajador
        ws = wb.create_sheet(title=trabajador.name)

        # Establecer los encabezados de la hoja
        ws.append(['Fecha', 'Ingreso', 'Salida', 'Refrigerio/Almuerzo', 'Hora Total'])

        # Obtener las asistencias de ese trabajador
        asistencias = Asistencia.objects.filter(trabajador=trabajador)

        # Poblar las filas con las asistencias
        for asistencia in asistencias:
            # Extraer datos
            fecha = asistencia.fecha
            hora_entrada = asistencia.hora if asistencia.tipo == 'entrada' else None
            hora_salida = asistencia.hora if asistencia.tipo == 'salida' else None
            # Asumimos que si no es salida, se calcula como refrigerio o almuerzo
            refrigerio = ''  # Aquí podrías agregar más lógica según tu modelo o datos
            hora_total = None

            # Si tenemos hora de entrada y salida, calculamos la duración
            if hora_entrada and hora_salida:
                # Calcular la diferencia de horas trabajadas (suponemos 1 hora de almuerzo por ejemplo)
                diferencia = datetime.combine(fecha, hora_salida) - datetime.combine(fecha, hora_entrada)
                hora_total = diferencia - timedelta(hours=1)  # Si hay 1 hora de almuerzo

            # Añadir los datos de la fila
            ws.append([fecha, hora_entrada, hora_salida, refrigerio, hora_total])

        # Autoajustar las columnas
        for col in range(1, 6):
            max_length = 0
            column = get_column_letter(col)
            for row in ws.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    # Configurar la respuesta para enviar el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=asistencias_trabajadores.xlsx'
    wb.save(response)
    return response

def registrar_asistencia(request):
    if request.method == 'POST':
        form = AsistenciaForm(request.POST, request.FILES)

        if form.is_valid():
            asistencia = form.save(commit=False)

            # Obtener el último registro del trabajador en el mismo día
            ultimo_registro = Asistencia.objects.filter(
                trabajador=asistencia.trabajador,
            ).order_by('-id').first()  # Obtenemos el último registro, si existe

            # Si hay un último registro
            if ultimo_registro:
                if ultimo_registro.tipo==asistencia.tipo:
                    error="Si aún no has registrado tu salida no puedes marcar una nueva entrada. Tampoco marcar 2 salidas seguidas, último registro: "+str(ultimo_registro.fecha)+" "+str(ultimo_registro.hora)+" "+ultimo_registro.tipo
                    messages.error(request,error)
                    return redirect('registrar_asistencia')
            # Si no hay un último registro, significa que es el primer registro del día
            elif asistencia.tipo == 'salida':
                messages.error(request, 'No puedes registrar una salida sin haber registrado previamente tu entrada.')
                return redirect('registrar_asistencia')

            # Si la validación es exitosa, guardamos la asistencia
            asistencia.save()
            messages.success(request, '¡Asistencia registrada con éxito!')
            return redirect('registrar_asistencia')
    else:
        form = AsistenciaForm()

    return render(request, 'asistencia/registrar_asistencia.html', {'form': form})

#Admin dashboard

@login_required
def vista_admin(request):
    # Vista para el admin que puede ver las asistencias de los trabajadores
    asistencias = Asistencia.objects.all()
    return render(request, 'asistencia/admin/admin_dashboard.html', {'asistencias': asistencias})

@method_decorator(login_required, name='dispatch')
class TrabajadorListView(ListView):
    model = Trabajador
    template_name = 'asistencia/admin/trabajador_list.html'
    context_object_name = 'trabajadores'

@method_decorator(login_required, name='dispatch')
class TrabajadorCreateView(CreateView):
    model = Trabajador
    template_name = 'asistencia/admin/trabajador_form.html'
    fields = ['name']  # Incluye aquí todos los campos que deseas manejar
    success_url = reverse_lazy('trabajador_list')

@method_decorator(login_required, name='dispatch')
class TrabajadorUpdateView(UpdateView):
    model = Trabajador
    template_name = 'asistencia/admin/trabajador_form.html'
    fields = ['name']  # Igual que en CreateView
    success_url = reverse_lazy('trabajador_list')

@method_decorator(login_required, name='dispatch')
class TrabajadorDeleteView(DeleteView):
    model = Trabajador
    template_name = 'asistencia/admin/trabajador_confirm_delete.html'
    success_url = reverse_lazy('trabajador_list')
    

# Vista para el Dashboard
@login_required
def dashboard(request):
    return render(request, 'dashboard/dashboard.html')

# Vista para gestionar asistencias (solo ejemplo)
@login_required
def gestionar_asistencias(request):
    # Aquí agregarías la lógica para gestionar asistencias
    return render(request, 'dashboard/gestionar_asistencias.html')

# Vista para gestionar errores (solo ejemplo)
@login_required
def gestionar_errores(request):
    # Aquí agregarías la lógica para gestionar errores
    return render(request, 'dashboard/gestionar_errores.html')


